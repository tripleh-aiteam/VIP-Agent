"""
realty_kb_loader — load Triple H's real property portfolio from Excel.

Reads `data/tripleh_properties.xlsx` (workbook with 18 sheets covering
apartments, commercial, land, factory units) and produces a structured
list of property entries the chatbot can ground LLM answers in.

The parser is intentionally LENIENT — each sheet has slightly different
column layouts, so we read by header name rather than fixed index and
fall back gracefully when a column is missing.

Public API:
  load_real_listings() -> list[dict]    — cached parsed listings
  service_areas()      -> list[str]     — derived list of regions
  reload()             -> None          — clear the cache (use after upload)

The chatbot service calls `load_real_listings()` and uses the result in
`_build_realty_system_prompt`. Cache TTL: 1 hour (refreshed on uvicorn
restart anyway).
"""

from __future__ import annotations

import os
import time
from pathlib import Path
from typing import Any

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover — runtime guard, not a build-time error
    openpyxl = None  # type: ignore[assignment]

from services.logger import log


# ----------------------------------------------------------------------------
#  Paths + cache
# ----------------------------------------------------------------------------

_DATA_DIR = Path(__file__).resolve().parent.parent / "data"
_WORKBOOK_PATH = _DATA_DIR / "tripleh_properties.xlsx"

_CACHE: dict[str, Any] = {"loaded_at": 0.0, "listings": None, "areas": None}
_CACHE_TTL_SECONDS = 3600  # 1 hour


# ----------------------------------------------------------------------------
#  Header normalization — different sheets use slightly different column names
# ----------------------------------------------------------------------------

# Map of canonical field → list of header substrings that should populate it.
# Order matters: we walk the actual header row and pick the FIRST match per
# canonical field. Substring match (case-sensitive Korean).
_FIELD_KEYWORDS: dict[str, tuple[str, ...]] = {
    "category":      ("구분",),                # 토지/아파트/상가/공장 etc.
    "unit_no":       ("호수", "순번"),         # unit number / sequence
    "address":       ("주소",),                # full address
    "use":           ("용도",),                # 공장용지/도로/임야 etc.
    "size_m2":       ("면적(㎡)", "공급면적", "면적 ㎡"),
    "size_pyeong":   ("면적(평)", "평", "면적 평"),
    "purchase_price":("분양가", "기존 매입가", "감정가격", "매입가"),
    "deposit":       ("보증금",),
    "rent":          ("월세",),
    "status":        ("현상태", "상태"),
    "contract":      ("계약기간", "임차기간"),
    "expiry":        ("계약 만료", "만료일"),
    "sale_plan":     ("매각 계획", "매각희망가", "매각예정가"),
    "yield_pre":     ("수익률 (세전)", "수익률(세전)", "수익률           (매각예정가 세전)"),
    "yield_post":    ("수익률 (세후)", "수익률(세후)"),
    "tenant":        ("임차인",),
    "tenant_phone":  ("전화번호", "연락처"),
    "tax":           ("세금", "재산세"),
    "notes":         ("비고", "메모", "매물 현황", "매물 동향", "물건 종합"),
    "official_value":("공시지가", "공시가격"),
}


def _normalize_header(text: str) -> str:
    """Collapse whitespace + newlines so 'size(평)\\n' matches '면적(평)'."""
    return " ".join((text or "").split())


def _detect_columns(header_row: list[str]) -> dict[str, int]:
    """Map canonical field → column index by matching header keywords.
    Returns {field: col_idx, …}. Fields whose keyword isn't found are absent."""
    out: dict[str, int] = {}
    normalized = [_normalize_header(str(h or "")) for h in header_row]
    for field, keywords in _FIELD_KEYWORDS.items():
        for idx, h in enumerate(normalized):
            if not h:
                continue
            if any(kw in h for kw in keywords):
                out[field] = idx
                break
    return out


def _stringify(val: Any) -> str:
    """Trim + sanitize a cell value for the LLM context (never None)."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    if isinstance(val, float):
        # Avoid scientific notation; round to integer KRW
        return f"{val:,.0f}" if abs(val) >= 10000 else str(val)
    s = str(val).strip()
    # Replace embedded newlines with spaces for one-line listing
    return " ".join(s.split())


def _is_data_row(row: list[Any]) -> bool:
    """Heuristic: a row is property data if it has at least 3 non-empty
    cells. Skips '합계' / 'TOTAL' rows."""
    non_empty = [c for c in row if c not in (None, "")]
    if len(non_empty) < 3:
        return False
    first_str = str(non_empty[0]).strip()
    if first_str.startswith(("합계", "총계", "총합")):
        return False
    return True


# ----------------------------------------------------------------------------
#  Sheet parsing
# ----------------------------------------------------------------------------

def _parse_sheet(ws: Any, sheet_name: str) -> list[dict[str, Any]]:
    """Extract listings from a single sheet. Header is typically at R5
    (after a '상세현황' label at R4)."""
    listings: list[dict[str, Any]] = []

    # Find the title (usually R1, first non-empty cell)
    title = ""
    for cell in ws[1]:
        if cell.value:
            title = str(cell.value).strip()
            break

    # Find header row — scan rows 3-7 for one that contains "구분"
    header_row_idx = 0
    header_values: list[str] = []
    for r in range(3, min(10, ws.max_row + 1)):
        row = [c.value for c in ws[r]]
        norm = [_normalize_header(str(v or "")) for v in row]
        if any("구분" in n for n in norm):
            header_row_idx = r
            header_values = row
            break
    if not header_row_idx:
        return listings

    col_map = _detect_columns(header_values)
    if not col_map:
        return listings

    # Data rows start after the header
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        if not _is_data_row(row):
            continue

        entry: dict[str, Any] = {"sheet": sheet_name, "title": title}
        for field, col_idx in col_map.items():
            if col_idx < len(row):
                entry[field] = _stringify(row[col_idx])

        # Skip rows with no useful info
        if not any(entry.get(k) for k in ("address", "unit_no", "category")):
            continue
        # Skip notes-only rows (e.g. "**공시지가 기준")
        if entry.get("category", "").startswith("**"):
            continue

        listings.append(entry)

    return listings


# ----------------------------------------------------------------------------
#  Workbook loading
# ----------------------------------------------------------------------------

def _load_workbook_now() -> tuple[list[dict[str, Any]], list[str]]:
    """Read the workbook from disk and parse all sheets. Returns
    (listings, service_areas)."""
    if openpyxl is None:
        log.warning("realty_kb_loader: openpyxl not installed — returning empty KB")
        return [], []

    if not _WORKBOOK_PATH.exists():
        log.warning(f"realty_kb_loader: workbook not found at {_WORKBOOK_PATH}")
        return [], []

    try:
        # NOTE: read_only=False intentionally. read_only mode breaks max_row
        # on some sheets (openpyxl bug — TypeError NoneType + int). The
        # workbook is small enough that loading fully is fine.
        wb = openpyxl.load_workbook(_WORKBOOK_PATH, data_only=True)
    except Exception as e:
        log.warning(f"realty_kb_loader: failed to open workbook: {e}")
        return [], []

    listings: list[dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        # Skip the summary sheet — it duplicates info that's in the per-property sheets
        if sheet_name.strip() == "총괄":
            continue
        try:
            ws = wb[sheet_name]
            parsed = _parse_sheet(ws, sheet_name)
            listings.extend(parsed)
        except Exception as e:
            log.warning(f"realty_kb_loader: failed to parse sheet '{sheet_name}': {e}")

    # Derive service areas from sheet titles + addresses
    area_keywords = ["파주", "향남", "의정부", "고척동", "보성리", "낙하리"]
    found_areas: set[str] = set()
    for entry in listings:
        haystack = (entry.get("title", "") + " " + entry.get("address", ""))
        for kw in area_keywords:
            if kw in haystack:
                found_areas.add(kw)

    log.info(
        f"realty_kb_loader: loaded {len(listings)} listings across "
        f"{len(found_areas)} areas",
        extra={"action": "realty_kb.loaded"},
    )
    return listings, sorted(found_areas)


def load_real_listings() -> list[dict[str, Any]]:
    """Return cached listings, refreshing from disk every hour."""
    now = time.time()
    if (
        _CACHE["listings"] is not None
        and now - _CACHE["loaded_at"] < _CACHE_TTL_SECONDS
    ):
        return _CACHE["listings"]
    listings, areas = _load_workbook_now()
    _CACHE["listings"] = listings
    _CACHE["areas"] = areas
    _CACHE["loaded_at"] = now
    return listings


def service_areas() -> list[str]:
    """Return derived service area names. Triggers load if cache empty."""
    load_real_listings()
    return _CACHE.get("areas") or []


def reload() -> None:
    """Force the cache to refresh on next call (e.g. after Excel upload)."""
    _CACHE["loaded_at"] = 0.0
    _CACHE["listings"] = None
    _CACHE["areas"] = None

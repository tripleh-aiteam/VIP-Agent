"""
Chatbot PERCEPTION pillar — converts arbitrary input (images, PDFs, Excel,
CSV, audio, sensor data) into TEXT that the TALK engine can reason about.

Flow:
  user uploads file → /chatbot/perceive → this module extracts text →
    chatbot.config sends the text to /chatbot/talk along with the user's question.

Each handler returns a dict { content: str, kind: str, meta: {...} }
where `content` is plain text the LLM can consume.
"""

from __future__ import annotations

import base64
import csv
import io
import os
from typing import Any

import httpx


# ---------------------------------------------------------------------------
# Image — Gemini Vision describes the image
# ---------------------------------------------------------------------------

async def perceive_image(file_bytes: bytes, content_type: str, user_hint: str = "") -> dict[str, Any]:
    """
    Send the image to Gemini Vision and return a textual description.
    `user_hint` is what the user typed alongside the image — helps Gemini
    focus its description on what the user actually wants to know.
    """
    gemini_key = os.getenv("GEMINI_API_KEY", "")
    if not gemini_key:
        return {"content": "[Image attached, but Gemini Vision is not configured]",
                "kind": "image", "meta": {"size_bytes": len(file_bytes)}}

    mime = content_type if content_type.startswith("image/") else "image/jpeg"
    b64 = base64.b64encode(file_bytes).decode("ascii")

    prompt = (
        f"Describe this image in detail. The user asked: \"{user_hint}\"\n\n"
        if user_hint else
        "Describe this image in detail — what's in it, any text, important objects, layout.\n\n"
    )
    prompt += (
        "If the image contains a chart or graph, describe its trends and key numbers. "
        "If it contains text, transcribe the text. "
        "If it's a diagram or UI screenshot, describe the structure and labels. "
        "Be concise but complete — output 2–6 sentences of plain text, no markdown."
    )

    body = {
        "contents": [{
            "parts": [
                {"inlineData": {"mimeType": mime, "data": b64}},
                {"text": prompt},
            ]
        }],
        "generationConfig": {"temperature": 0.2, "maxOutputTokens": 600},
    }

    try:
        async with httpx.AsyncClient(timeout=60) as client:
            r = await client.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}",
                json=body,
            )
            if r.status_code != 200:
                return {"content": f"[Vision error: HTTP {r.status_code}]", "kind": "image",
                        "meta": {"error": r.text[:200]}}
            j = r.json()
            text = j["candidates"][0]["content"]["parts"][0]["text"].strip()
            return {"content": text, "kind": "image", "meta": {"size_bytes": len(file_bytes), "engine": "gemini-2.5-flash"}}
    except Exception as e:
        return {"content": f"[Vision call failed: {e}]", "kind": "image", "meta": {"error": str(e)}}


# ---------------------------------------------------------------------------
# PDF — extract text from each page
# ---------------------------------------------------------------------------

def perceive_pdf(file_bytes: bytes) -> dict[str, Any]:
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        pages: list[str] = []
        for i, page in enumerate(reader.pages[:50]):  # cap at 50 pages
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"[Page {i + 1}]\n{text.strip()}")
        full = "\n\n".join(pages)
        if len(full) > 12_000:
            full = full[:12_000] + f"\n\n[...truncated. Total {len(reader.pages)} pages, showing first ~12K chars]"
        return {"content": full or "[PDF appears empty or image-only]",
                "kind": "pdf",
                "meta": {"page_count": len(reader.pages), "size_bytes": len(file_bytes)}}
    except Exception as e:
        return {"content": f"[PDF parse failed: {e}]", "kind": "pdf", "meta": {"error": str(e)}}


# ---------------------------------------------------------------------------
# Excel — describe sheets + sample rows
# ---------------------------------------------------------------------------

def perceive_excel(file_bytes: bytes) -> dict[str, Any]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        out: list[str] = []
        total_rows = 0
        for sheet_name in wb.sheetnames[:8]:  # cap at 8 sheets
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            total_rows += len(rows)
            if not rows:
                out.append(f"[Sheet: {sheet_name}] empty")
                continue
            headers = [str(c) if c is not None else "" for c in (rows[0] or [])]
            out.append(f"[Sheet: {sheet_name}] {len(rows)} rows")
            out.append(f"  Columns: {', '.join(headers)}")
            # Sample first 5 data rows
            sample = rows[1:6]
            for ri, r in enumerate(sample):
                vals = [str(c) if c is not None else "" for c in r]
                out.append(f"  Row {ri + 2}: " + " | ".join(vals[:8]))
            if len(rows) > 6:
                out.append(f"  ... and {len(rows) - 6} more rows")
        return {"content": "\n".join(out) or "[Empty workbook]",
                "kind": "excel",
                "meta": {"sheet_count": len(wb.sheetnames), "total_rows": total_rows, "size_bytes": len(file_bytes)}}
    except Exception as e:
        return {"content": f"[Excel parse failed: {e}]", "kind": "excel", "meta": {"error": str(e)}}


# ---------------------------------------------------------------------------
# CSV — describe headers + sample
# ---------------------------------------------------------------------------

def perceive_csv(file_bytes: bytes) -> dict[str, Any]:
    try:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        if not rows:
            return {"content": "[Empty CSV]", "kind": "csv", "meta": {"size_bytes": len(file_bytes)}}
        headers = rows[0]
        out = [f"[CSV — {len(rows)} rows]",
               f"  Columns: {', '.join(headers)}"]
        for ri, r in enumerate(rows[1:6]):
            out.append(f"  Row {ri + 2}: " + " | ".join(r[:8]))
        if len(rows) > 6:
            out.append(f"  ... and {len(rows) - 6} more rows")
        return {"content": "\n".join(out),
                "kind": "csv",
                "meta": {"row_count": len(rows), "size_bytes": len(file_bytes)}}
    except Exception as e:
        return {"content": f"[CSV parse failed: {e}]", "kind": "csv", "meta": {"error": str(e)}}


# ---------------------------------------------------------------------------
# DOCX — extract text
# ---------------------------------------------------------------------------

def perceive_docx(file_bytes: bytes) -> dict[str, Any]:
    try:
        import docx
        d = docx.Document(io.BytesIO(file_bytes))
        paras = [p.text for p in d.paragraphs if p.text.strip()]
        text = "\n".join(paras)
        if len(text) > 12_000:
            text = text[:12_000] + "\n[...truncated]"
        return {"content": text or "[DOCX appears empty]",
                "kind": "docx",
                "meta": {"paragraph_count": len(paras), "size_bytes": len(file_bytes)}}
    except Exception as e:
        return {"content": f"[DOCX parse failed: {e}]", "kind": "docx", "meta": {"error": str(e)}}


# ---------------------------------------------------------------------------
# Plain text
# ---------------------------------------------------------------------------

def perceive_text(file_bytes: bytes) -> dict[str, Any]:
    try:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        if len(text) > 12_000:
            text = text[:12_000] + "\n[...truncated]"
        return {"content": text, "kind": "text", "meta": {"size_bytes": len(file_bytes)}}
    except Exception as e:
        return {"content": f"[Text decode failed: {e}]", "kind": "text", "meta": {"error": str(e)}}


# ---------------------------------------------------------------------------
# Dispatcher — pick handler from MIME / filename
# ---------------------------------------------------------------------------

async def perceive_file(filename: str, content_type: str, file_bytes: bytes,
                        user_hint: str = "") -> dict[str, Any]:
    """Route to the right handler based on file extension / mime type."""
    ct = (content_type or "").lower()
    name = (filename or "").lower()
    ext = name.rsplit(".", 1)[-1] if "." in name else ""

    if ct.startswith("image/") or ext in ("png", "jpg", "jpeg", "gif", "webp", "bmp", "heic"):
        return await perceive_image(file_bytes, content_type or "image/jpeg", user_hint)

    if ct in ("application/pdf",) or ext == "pdf":
        return perceive_pdf(file_bytes)

    if ct in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
              "application/vnd.ms-excel") or ext in ("xlsx", "xlsm", "xls"):
        return perceive_excel(file_bytes)

    if ct in ("text/csv", "application/csv") or ext == "csv":
        return perceive_csv(file_bytes)

    if ct in ("application/vnd.openxmlformats-officedocument.wordprocessingml.document",) or ext in ("docx", "doc"):
        return perceive_docx(file_bytes)

    if ct.startswith("text/") or ext in ("txt", "md", "json", "log"):
        return perceive_text(file_bytes)

    return {"content": f"[Unsupported file type: {content_type or ext or 'unknown'}]",
            "kind": "unknown",
            "meta": {"size_bytes": len(file_bytes), "filename": filename, "content_type": content_type}}

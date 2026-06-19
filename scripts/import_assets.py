"""
import_assets.py — load the company asset workbook into Supabase.

Parses the multi-sheet 자산관리 Excel into two tables the orchestrator reads:

  asset_portfolio  ← the 총괄 (summary) sheet: one row per portfolio line item
                     (category, description, value, deposit, monthly_rent).
  asset_units      ← each property sheet's per-unit/per-parcel table, normalized
                     to common columns + a JSONB `extra` for type-specific fields.

Re-runnable: it CREATE TABLE IF NOT EXISTS, then full-refreshes both tables, so
running it again after editing the Excel just reloads the latest data.

Usage (from vip-ai-platform/):
    python scripts/import_assets.py "C:/path/to/자산관리.xlsx"
(defaults to the workbook on the Desktop if no path given). DATABASE_URL is read
from .env.supabase / .env so it writes to the SAME DB the orchestrator reads.
"""

from __future__ import annotations

import json
import os
import re
import sys

DEFAULT_XLSX = r"C:\Users\TRIPLEH\Desktop\VIP Agent\자산관리_ver.1_260206 (2).xlsx"


def _load_env() -> None:
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # vip-ai-platform/
    for fn in (".env.supabase", ".env"):
        p = os.path.join(here, fn)
        if not os.path.exists(p):
            continue
        for ln in open(p, encoding="utf-8"):
            ln = ln.strip()
            if ln and not ln.startswith("#") and "=" in ln:
                k, v = ln.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def _num(v):
    """Parse a cell to float, or None for blank / #REF! / text."""
    if v is None:
        return None
    s = str(v).replace(",", "").strip()
    if not s or s.upper().startswith(("#REF", "#VALUE", "#DIV", "#N/A")):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return float(m.group(0)) if m else None


def _norm(h) -> str:
    """Normalize a header cell: drop whitespace/newlines for keyword matching."""
    return re.sub(r"\s+", "", str(h or ""))


# header-keyword -> canonical column. Order matters (checked top-down).
def _canon(header: str) -> str | None:
    h = _norm(header)
    if not h:
        return None
    if "순번" in h or h in ("NO", "NO."):
        return "seq"
    if h == "구분":
        return "category"
    if "호수" in h or "호실" in h:
        return "unit_no"
    if "주소" in h or "위치" in h or "소재지" in h or h == "BL":
        return "address"
    if "평" in h and "면적" in h:          # 면적(평) / 전용면적(평) / 분양면적(평)
        return "area_pyeong"
    if "면적" in h or "㎡" in h:            # 공급면적(㎡) / 면적(㎡)
        return "area_m2"
    if "현시세" in h or (h.startswith("시세")):
        return "market_value"
    if "공시" in h:                         # 공시가격 / 공시지가 / 2024공시가격
        return "official_price"
    if "분양가" in h or "매입가" in h or "감정가" in h or "분양예정가" in h \
            or "자산가치" in h or h == "원가" or "분양예정" in h:
        return "price"
    if "보증금" in h:
        return "deposit"
    if "월세" in h:
        return "monthly_rent"
    if "임차인" in h:
        return "tenant"
    if "현상태" in h or h == "상태" or "영업현황" in h or "분양여부" in h:
        return "status"
    if h == "내용":
        return "description"
    return None


CANON_FIELDS = ("seq", "category", "unit_no", "address", "area_m2", "area_pyeong",
                "price", "official_price", "market_value", "deposit", "monthly_rent",
                "tenant", "status", "description")
HDR_HINTS = ("구분", "순번", "호수", "호실", "주소", "위치", "면적", "분양가", "매입가",
             "감정가", "보증금", "월세", "공시", "현시세", "상태", "용도", "분양여부")


def _find_header_row(ws, scan=12):
    best = None
    for r in range(1, min(ws.max_row, scan) + 1):
        vals = [_norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
        hits = sum(1 for v in vals for k in HDR_HINTS if k and k in v)
        if hits >= 3:
            return r
    return best


def parse_portfolio(ws) -> list[dict]:
    """총괄 → portfolio line items."""
    rows = []
    # header at the row containing 회사/구분/분양가 ; data starts next row
    hr = _find_header_row(ws)
    if not hr:
        return rows
    # locate columns by header text
    colmap = {}
    for c in range(1, ws.max_column + 1):
        h = _norm(ws.cell(hr, c).value)
        if "구분" in h:
            colmap["category"] = c
        elif h == "내용":
            colmap["description"] = c
        elif "분양가" in h or "자산가치" in h:
            colmap["price"] = c
        elif "보증금" in h:
            colmap["deposit"] = c
        elif "월세" in h:
            colmap["monthly_rent"] = c
        elif "순번" in h or h in ("NO", "NO."):
            colmap["seq"] = c
    last_cat = None
    for r in range(hr + 1, ws.max_row + 1):
        cat = ws.cell(r, colmap.get("category", 3)).value if colmap.get("category") else None
        desc = ws.cell(r, colmap.get("description", 4)).value if colmap.get("description") else None
        cat = (str(cat).strip() if cat not in (None, "") else None) or last_cat
        if cat:
            last_cat = cat
        if not desc or not str(desc).strip():
            continue
        if "합계" in str(desc) or "합계" in str(ws.cell(r, colmap.get("seq", 2)).value or ""):
            continue
        rows.append({
            "seq": int(_num(ws.cell(r, colmap["seq"]).value)) if colmap.get("seq") and _num(ws.cell(r, colmap["seq"]).value) else None,
            "category": cat,
            "description": str(desc).strip(),
            "sale_price": _num(ws.cell(r, colmap["price"]).value) if colmap.get("price") else None,
            "deposit": _num(ws.cell(r, colmap["deposit"]).value) if colmap.get("deposit") else None,
            "monthly_rent": _num(ws.cell(r, colmap["monthly_rent"]).value) if colmap.get("monthly_rent") else None,
        })
    return rows


def parse_units(ws, sheet_name: str) -> list[dict]:
    """A property sheet → normalized unit rows (+ extras JSON)."""
    hr = _find_header_row(ws)
    if not hr:
        return []
    # canonical col map (first match wins) + collect extras
    canon_col, extra_col = {}, {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(hr, c).value
        cn = _canon(raw)
        label = re.sub(r"\s+", " ", str(raw or "")).strip()
        if cn and cn not in canon_col:
            canon_col[cn] = c
        elif label:
            extra_col[label] = c
    rows = []
    last_cat = None
    blanks = 0
    for r in range(hr + 1, ws.max_row + 1):
        # qualify as a data row: numeric seq OR a non-empty unit_no/address
        seq_v = _num(ws.cell(r, canon_col["seq"]).value) if canon_col.get("seq") else None
        unit_v = ws.cell(r, canon_col["unit_no"]).value if canon_col.get("unit_no") else None
        addr_v = ws.cell(r, canon_col["address"]).value if canon_col.get("address") else None
        if seq_v is None and not (unit_v and str(unit_v).strip()) and not (addr_v and str(addr_v).strip()):
            blanks += 1
            if blanks >= 3:
                break          # likely past the table / into a sub-section
            continue
        blanks = 0
        cat = ws.cell(r, canon_col["category"]).value if canon_col.get("category") else None
        cat = (str(cat).strip() if cat not in (None, "") else None) or last_cat
        if cat:
            last_cat = cat
        rec = {"property": sheet_name, "category": cat, "source_sheet": sheet_name}
        for f in ("unit_no", "address", "tenant", "status"):
            if canon_col.get(f):
                v = ws.cell(r, canon_col[f]).value
                rec[f] = str(v).strip() if v not in (None, "") else None
        for f in ("area_m2", "area_pyeong", "price", "official_price", "market_value", "deposit", "monthly_rent"):
            if canon_col.get(f):
                rec[f] = _num(ws.cell(r, canon_col[f]).value)
        extra = {}
        for label, c in extra_col.items():
            v = ws.cell(r, c).value
            if v not in (None, ""):
                extra[label] = (_num(v) if _num(v) is not None else str(v).strip())
        rec["extra"] = json.dumps(extra, ensure_ascii=False) if extra else None
        # skip totally empty rows (only property set)
        if any(rec.get(k) is not None for k in ("unit_no", "address", "price", "monthly_rent", "deposit", "area_m2")):
            rows.append(rec)
    return rows


DDL = """
CREATE TABLE IF NOT EXISTS asset_portfolio (
    id           BIGSERIAL PRIMARY KEY,
    seq          INTEGER,
    category     TEXT,
    description  TEXT,
    sale_price   NUMERIC,
    deposit      NUMERIC,
    monthly_rent NUMERIC,
    source_sheet TEXT DEFAULT '총괄',
    imported_at  TIMESTAMPTZ DEFAULT now()
);
CREATE TABLE IF NOT EXISTS asset_units (
    id             BIGSERIAL PRIMARY KEY,
    property       TEXT,
    category       TEXT,
    unit_no        TEXT,
    address        TEXT,
    area_m2        NUMERIC,
    area_pyeong    NUMERIC,
    price          NUMERIC,
    official_price NUMERIC,
    market_value   NUMERIC,
    deposit        NUMERIC,
    monthly_rent   NUMERIC,
    status         TEXT,
    tenant         TEXT,
    extra          JSONB,
    source_sheet   TEXT,
    imported_at    TIMESTAMPTZ DEFAULT now()
);
"""


def main():
    _load_env()
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    import openpyxl
    from sqlalchemy import create_engine, text

    url = os.environ["DATABASE_URL"]
    eng = create_engine(url, pool_pre_ping=True)
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    portfolio = parse_portfolio(wb["총괄"]) if "총괄" in wb.sheetnames else []
    units, per_sheet = [], {}
    for name in wb.sheetnames:
        if name == "총괄":
            continue
        rs = parse_units(wb[name], name)
        if rs:
            units.extend(rs)
            per_sheet[name] = len(rs)

    with eng.begin() as c:
        for stmt in DDL.strip().split(";"):
            if stmt.strip():
                c.execute(text(stmt))
        c.execute(text("DELETE FROM asset_portfolio"))
        c.execute(text("DELETE FROM asset_units"))
        for p in portfolio:
            c.execute(text("""INSERT INTO asset_portfolio (seq,category,description,sale_price,deposit,monthly_rent)
                              VALUES (:seq,:category,:description,:sale_price,:deposit,:monthly_rent)"""), p)
        for u in units:
            cols = ["property", "category", "unit_no", "address", "area_m2", "area_pyeong",
                    "price", "official_price", "market_value", "deposit", "monthly_rent",
                    "status", "tenant", "extra", "source_sheet"]
            params = {k: u.get(k) for k in cols}
            c.execute(text(
                "INSERT INTO asset_units (" + ",".join(cols) + ") VALUES (" +
                ",".join(":" + k for k in cols) + ")"), params)

    print(f"portfolio rows: {len(portfolio)}")
    print(f"unit rows: {len(units)}  across {len(per_sheet)} sheets")
    for k, v in per_sheet.items():
        print(f"   {k}: {v}")
    # quick totals readback
    with eng.connect() as c:
        pv = c.execute(text("SELECT COALESCE(SUM(sale_price),0), COALESCE(SUM(monthly_rent),0) FROM asset_portfolio")).first()
        uv = c.execute(text("SELECT COUNT(*), COALESCE(SUM(monthly_rent),0) FROM asset_units")).first()
    pval, prent = float(pv[0]), float(pv[1])
    print(f"\nasset_portfolio: total value {pval:,.0f}원 (~{pval/1e8:,.1f}억), monthly_rent {prent:,.0f}원")
    print(f"asset_units: {uv[0]} rows, monthly_rent {float(uv[1]):,.0f}원")


if __name__ == "__main__":
    main()

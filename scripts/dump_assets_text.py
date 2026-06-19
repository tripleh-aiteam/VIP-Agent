"""
dump_assets_text.py — flatten the whole asset workbook into one clean Markdown
text file, so it can be ingested into the chatbot knowledge base (RAG) and the
assistant can recall ANY cell / sub-table / note across every sheet.

Every non-empty sheet becomes a '## <sheet>' section with all non-empty rows
rendered as ' | '-joined cells (merged cells / sub-tables / Gantt labels all
included as plain text). #REF!/error cells are shown as '(오류)'.

Usage:  python scripts/dump_assets_text.py [xlsx_path] [out_md_path]
"""

from __future__ import annotations

import sys

DEFAULT_XLSX = r"C:\Users\TRIPLEH\Desktop\VIP Agent\자산관리_ver.1_260206 (2).xlsx"
DEFAULT_OUT = r"C:\Users\TRIPLEH\Desktop\VIP Agent\vip-ai-platform\data\asset_workbook_full.md"


def _cell(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().replace("\n", " ")
    if s.startswith("#") and s.upper().rstrip("!").rstrip() in ("#REF", "#VALUE", "#DIV/0", "#N/A", "#NAME?"):
        return "(오류)"
    return s


def main():
    import os
    import openpyxl
    xlsx = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    out = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUT
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    lines = ["# 트리플에이치 자산관리 워크북 (전체 시트 텍스트)",
             "이 문서는 회사 자산관리 엑셀의 모든 시트를 텍스트로 변환한 것입니다. "
             "각 시트의 모든 행/열, 하위 표, 일정, 메모를 포함합니다.\n"]
    sheets_done = 0
    for name in wb.sheetnames:
        ws = wb[name]
        rows_txt = []
        for row in ws.iter_rows():
            cells = [_cell(c.value) for c in row]
            # trim trailing empties
            while cells and cells[-1] == "":
                cells.pop()
            if any(x for x in cells):
                rows_txt.append(" | ".join(cells))
        if not rows_txt:
            continue
        sheets_done += 1
        lines.append(f"\n## 시트: {name}\n")
        lines.extend(rows_txt)
    text = "\n".join(lines)
    with open(out, "w", encoding="utf-8") as f:
        f.write(text)
    print(f"wrote {out}")
    print(f"sheets with data: {sheets_done} | chars: {len(text):,} | lines: {text.count(chr(10)):,}")


if __name__ == "__main__":
    main()

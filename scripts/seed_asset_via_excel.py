"""
Alternate seed path: build an Excel file matching the Asset Agent's
'/api/upload/template' format and POST it to '/api/upload/excel'.

This path bypasses the buggy '/api/manage/properties' endpoint that has a
known UniqueViolation on the global property_id sequence. The Excel uploader
generates codes via a different code path and may handle conflicts properly.

Sheets the backend expects (Korean):
  자산   (Properties / Buildings)
  호실   (Units)
  임차인 (Tenants)
  임대계약 (Lease Contracts)
"""

from __future__ import annotations

import os
import sys
from datetime import date, timedelta
from pathlib import Path

import httpx
import openpyxl

_REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO_ROOT))
try:
    from dotenv import load_dotenv
    load_dotenv(_REPO_ROOT / ".env", override=False)
except ImportError:
    pass

from adapters.mock_data import ASSETS  # noqa: E402

ASSET_URL = os.getenv("REAL_ASSET_AGENT_URL", "https://asset-agent-s4tw.onrender.com")
EMAIL = os.getenv("ASSET_AGENT_EMAIL", "vip-orchestrator@tripleh.com")
PASSWORD = os.getenv("ASSET_AGENT_PASSWORD", "VipAgent2026!")
OUT_XLSX = _REPO_ROOT / "data" / "uploads" / "asset" / "vip_seed.xlsx"
OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

NAME_TO_DISTRICT = {
    "Gangnam Office Tower":   ("강남구",   "서울시 강남구 테헤란로 123"),
    "Yeouido Plaza":          ("영등포구", "서울시 영등포구 여의도동 100"),
    "Pangyo Tech Center":     ("분당구",   "경기 성남시 분당구 판교역로 200"),
    "Jamsil Retail Building": ("송파구",   "서울시 송파구 올림픽로 88"),
    "Itaewon Hotel":          ("용산구",   "서울시 용산구 이태원로 50"),
    "Songdo Logistics Hub":   ("연수구",   "인천 연수구 송도국제도시"),
    "Hongdae Apartment Bldg": ("마포구",   "서울시 마포구 홍익로 30"),
    "Busan Marine Tower":     ("해운대구", "부산 해운대구 마린시티 45"),
}

TYPE_KO = {
    "office":      "오피스",
    "retail":      "상가",
    "hospitality": "호텔",
    "industrial":  "물류",
    "residential": "오피스텔",
}

TENANTS_KO = [
    ("김민준", "삼성생명",       "010-1234-5678", "kim@samsunglife.kr",  "보험"),
    ("이지훈", "LG CNS",         "010-2345-6789", "lee@lgcns.com",       "IT"),
    ("박서연", "네이버 클라우드", "010-3456-7890", "park@navercloud.com", "IT"),
    ("최현우", "롯데백화점",     "010-4567-8901", "choi@lotte.kr",       "유통"),
    ("정민서", "현대호텔",       "010-5678-9012", "jung@hyundaihotel.com","호텔"),
    ("한수빈", "CJ 대한통운",    "010-6789-0123", "han@cjlogistics.com", "물류"),
    ("윤태훈", "카카오모빌리티", "010-7890-1234", "yoon@kakaomobility.com","IT"),
    ("신은지", "부산해양산업",   "010-8901-2345", "shin@bmind.kr",       "조선"),
]


def build_excel(path: Path) -> None:
    wb = openpyxl.Workbook()
    # Default sheet → 자산
    ws_p = wb.active
    ws_p.title = "자산"
    p_headers = ["건물명", "주소", "구/군", "유형", "총면적(㎡)", "총면적(평)", "임대면적(㎡)",
                 "층수", "건물연식(년)", "취득가(원)", "취득일", "공시지가(원)", "용도지역",
                 "현상태", "세금(원)", "수익률(세전)", "수익률(세후)", "매각희망가(원)",
                 "대출담보액(원)", "매물동향", "매각계획", "비고"]
    ws_p.append(p_headers)

    ws_u = wb.create_sheet("호실")
    u_headers = ["건물명", "층", "호실번호", "면적(㎡)", "면적(평)", "유형", "분양가(원)",
                 "공시가격(원)", "현상태", "수익률(세전)", "수익률(세후)", "매물동향",
                 "물건종합", "비고"]
    ws_u.append(u_headers)

    ws_t = wb.create_sheet("임차인")
    t_headers = ["성명", "상호명", "상호명/업종", "사업자번호", "전화번호", "이메일",
                 "업종", "임차건물", "호실", "임차신고여부", "비고"]
    ws_t.append(t_headers)

    ws_l = wb.create_sheet("임대계약")
    l_headers = ["건물명", "호실번호", "임차인명", "계약시작일", "계약종료일", "월임대료(원)",
                 "관리비(원)", "보증금(원)", "수익률(세전)", "수익률(세후)", "세금(원)",
                 "미수금(원)", "미납내역", "임대료납부일", "비고"]
    ws_l.append(l_headers)

    start = (date.today() - timedelta(days=180)).isoformat()
    end   = (date.today() + timedelta(days=550)).isoformat()

    for i, asset in enumerate(ASSETS):
        district, address = NAME_TO_DISTRICT.get(asset["name"], ("강남구", "서울시"))
        bldg_name_ko = asset["name"]  # keep English name; backend just stores string
        type_ko = TYPE_KO.get(asset["type"], "오피스")
        area_sqm = round(asset["value_krw"] / 12_000_000, 0)
        area_pyeong = round(area_sqm * 0.3025, 1)

        # 자산 sheet
        ws_p.append([
            bldg_name_ko, address, district, type_ko,
            area_sqm, area_pyeong, area_sqm * 0.9,
            max(3, int(area_sqm / 200)), 8,
            int(asset["value_krw"] * 0.85), "2018-06-15",
            int(asset["value_krw"] * 0.6), "상업지역",
            None, None, None, None, None, None, None, None, None,
        ])

        # 호실 — single unit
        ws_u.append([
            bldg_name_ko, 1, "101호",
            area_sqm * 0.9, round(area_sqm * 0.9 * 0.3025, 1),
            type_ko, None, None, "임대중", None, None, None, None, None,
        ])

        # 임차인
        name, biz, phone, email, industry = TENANTS_KO[i % len(TENANTS_KO)]
        ws_t.append([
            name, biz, biz, f"123-45-{6789 + i:05d}", phone, email,
            industry, bldg_name_ko, "101호", None, None,
        ])

        # 임대계약
        monthly = asset["monthly_income_krw"]
        ws_l.append([
            bldg_name_ko, "101호", name, start, end,
            monthly, int(monthly * 0.1), monthly * 12,
            None, None, None, None, None, None, None,
        ])

    wb.save(path)


def login(client: httpx.Client) -> str:
    r = client.post(f"{ASSET_URL}/api/auth/login",
                    json={"email": EMAIL, "password": PASSWORD},
                    timeout=15)
    r.raise_for_status()
    return r.json()["access_token"]


def upload_excel(client: httpx.Client, token: str, path: Path) -> dict:
    headers = {"Authorization": f"Bearer {token}"}
    with open(path, "rb") as f:
        files = {"file": (path.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = client.post(f"{ASSET_URL}/api/upload/excel", headers=headers, files=files, timeout=120)
    print(f"Upload status: {r.status_code}")
    try:
        body = r.json()
        return body
    except Exception:
        return {"raw": r.text[:500]}


def main() -> int:
    print(f"Building Excel at: {OUT_XLSX}")
    build_excel(OUT_XLSX)
    print(f"Excel ready ({OUT_XLSX.stat().st_size} bytes), {len(ASSETS)} properties.")
    print()

    with httpx.Client() as c:
        try:
            tok = login(c)
            print("Login OK.")
        except Exception as e:
            print(f"Login failed: {e}")
            return 1

        print("Uploading Excel to /api/upload/excel...")
        result = upload_excel(c, tok, OUT_XLSX)
        import json
        try:
            print(json.dumps(result, ensure_ascii=False, indent=2)[:2000])
        except Exception:
            print(result)

        # Verify
        print()
        print("Verifying dashboard summary...")
        r = c.get(f"{ASSET_URL}/api/dashboard/summary",
                  headers={"Authorization": f"Bearer {tok}"}, timeout=15)
        if r.status_code == 200:
            data = r.json().get("data", {})
            print(f"  total_properties: {data.get('total_properties', 0)}")
            print(f"  total_units:      {data.get('total_units', 0)}")
            print(f"  monthly_income:   {data.get('monthly_rental_income', 0):,} KRW")
        else:
            print(f"  dashboard fetch failed: HTTP {r.status_code}")
    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.exit(main())
